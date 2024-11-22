# T3Import.py
# T3Import.py： 负责将 T3.csv 中的结果数据导入到 Answer.xls 的 attachment1 表中，更新对应的 PSNR、UCIQE、UIQM 列。
import pandas as pd
from openpyxl import load_workbook

# 读取 T3.csv 文件
t3_df = pd.read_csv('T3.csv')

# 读取 Answer.xls 的 attachment1 表
answer_file_path = '../../results/metrics/Answer.xls'
answer_sheet_name = 'attachment1 results'

# 加载现有的 Excel 工作簿
book = load_workbook(answer_file_path)

# 读取 attachment1 表为 DataFrame
answer_df = pd.read_excel(answer_file_path, sheet_name=answer_sheet_name)

# 将 T3.csv 中的数据合并到 Answer.xls 的 attachment1 表中
# 根据 'image file name' 列进行匹配，更新 PSNR、UCIQE、UIQM 列
merged_df = pd.merge(answer_df, t3_df[['image file name', 'PSNR', 'UCIQE', 'UIQM']], on='image file name', how='left')

# 从工作簿中删除旧的 attachment1 表
if answer_sheet_name in book.sheetnames:
    std = book[answer_sheet_name]
    book.remove(std)
    book.save(answer_file_path)

# 将更新后的数据写回到 attachment1 表中
with pd.ExcelWriter(answer_file_path, engine='openpyxl', mode='a') as writer:
    writer.book = book
    merged_df.to_excel(writer, sheet_name=answer_sheet_name, index=False)
    writer.save()
